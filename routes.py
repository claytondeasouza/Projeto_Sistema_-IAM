# Importações existentes
from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_user, logout_user, login_required, current_user
from flask import current_app as app

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import re

# --- NOVAS IMPORTAÇÕES NECESSÁRIAS ---
import os
from datetime import datetime

# --- Modelos e DB (sem alteração) ---
from .models import User, USERS_DB


# ==============================================================================
# --- ETAPA 1: CONFIGURAÇÃO DA PASTA DE AUTOMAÇÃO (Ajuste Obrigatório) ---
#
# Defina o caminho completo para a pasta que será monitorada pelo Power Automate.
# O servidor onde esta aplicação Flask roda DEVE ter permissão de escrita nesta pasta.
#
# Exemplo para Windows (se a pasta estiver sincronizada via OneDrive/SharePoint):
PASTA_AUTOMACAO = r'C:\Users\clayton.souza\OneDrive - Think IT\Procedimentos Técnicos\projeto_gestão_acessos\formulario_solicitacao'
#
# Exemplo para servidor Linux:
# PASTA_AUTOMACAO = '/home/usuario/sharepoint/formularios_iam'
#
# ==============================================================================

# Verifica se a pasta de destino existe, se não, a cria.
if not os.path.exists(PASTA_AUTOMACAO):
    try:
        os.makedirs(PASTA_AUTOMACAO)
        print(f"Pasta de automação criada em: {PASTA_AUTOMACAO}")
    except OSError as e:
        print(f"ERRO: Não foi possível criar a pasta de destino: {e}")
        # Em um ambiente de produção, você pode querer tratar este erro de forma mais robusta.


# --- FUNÇÕES DE LOGIN/LOGOUT (sem alteração) ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('solicitacao'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_obj = User.find_by_username(username)
        if user_obj and USERS_DB[user_obj.id]['password'] == password:
            login_user(user_obj)
            return redirect(url_for('solicitacao'))
        else:
            flash('Usuário ou senha inválidos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu do sistema.', 'info')
    return redirect(url_for('login'))

# --- ROTA PRINCIPAL (sem alteração) ---
@app.route('/', methods=['GET', 'POST'])
@app.route('/solicitacao', methods=['GET', 'POST'])
@login_required
def solicitacao():
    if request.method == 'POST':
        # A lógica agora chama a função que salva o arquivo e redireciona
        return gerar_e_salvar_excel(request.form)
    
    return render_template('solicitacao.html')


# --- FUNÇÃO DE PARSE (sem alteração) ---
def parse_table_data(form, prefix, fields):
    items = []
    pattern = re.compile(f'^{prefix}\\[(\\d+)\\]\\[(.+?)\\]$')
    indexed_data = {}
    for key, value in form.items():
        match = pattern.match(key)
        if match:
            index = int(match.group(1))
            field_name = match.group(2)
            if index not in indexed_data:
                indexed_data[index] = {}
            indexed_data[index][field_name] = value
    for index in sorted(indexed_data.keys()):
        item = indexed_data[index]
        if any(item.get(field, '').strip() for field in fields):
            items.append(item)
    return items


# --- FUNÇÃO PRINCIPAL DE GERAÇÃO DE EXCEL (TOTALMENTE ATUALIZADA) ---
def gerar_e_salvar_excel(form):
    try:
        workbook = openpyxl.Workbook()
        
        # --- (Toda a lógica de criação de abas e estilos permanece a mesma) ---
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # ABA 1: DADOS GERAIS
        sheet_geral = workbook.active
        sheet_geral.title = "Dados Gerais"
        dados_gerais_fields = [
            'nome_completo', 'matricula', 'cargo', 'departamento', 'tipo_vinculo', 'nome_consultoria',
            'gestor_consultoria', 'usuario_espelho', 'gestor', 'data_solicitacao', 'data_demissao',
            'tipo_solicitacao', 'tipo_licenca'
        ]
        sheet_geral.append(["Campo", "Valor"])
        for field in dados_gerais_fields:
            valor = form.get(field, '')
            if valor: sheet_geral.append([field.replace("_", " ").title(), valor])
        sheet_geral.column_dimensions['A'].width = 30
        sheet_geral.column_dimensions['B'].width = 50

        # ABA 2: SERVIDORES
        servidores_fields = ['servidor', 'perfil_acesso', 'acao', 'observacoes']
        servidores_data = parse_table_data(form, 'servidores', servidores_fields)
        if servidores_data:
            sheet_servidores = workbook.create_sheet(title="Servidores")
            sheet_servidores.append([f.replace("_", " ").title() for f in servidores_fields])
            for row in servidores_data: sheet_servidores.append([row.get(f, '') for f in servidores_fields])
            for col in sheet_servidores.columns: sheet_servidores.column_dimensions[col[0].column_letter].width = 30
        
        # ... (Lógica para as abas 3, 4, 5, 6 permanece idêntica) ...
        # ABA 3: SISTEMAS
        sistemas_fields = ['sistema', 'perfil_acesso', 'acao', 'observacoes']
        sistemas_data = parse_table_data(form, 'sistemas', sistemas_fields)
        if sistemas_data:
            sheet_sistemas = workbook.create_sheet(title="Sistemas")
            sheet_sistemas.append([f.replace("_", " ").title() for f in sistemas_fields])
            for row in sistemas_data: sheet_sistemas.append([row.get(f, '') for f in sistemas_fields])
            for col in sheet_sistemas.columns: sheet_sistemas.column_dimensions[col[0].column_letter].width = 30
        
        # ABA 4: PASTAS DE REDE
        pastas_fields = ['pasta', 'tipo_acesso', 'acao', 'observacoes']
        pastas_data = parse_table_data(form, 'pastas_rede', pastas_fields)
        if pastas_data:
            sheet_pastas = workbook.create_sheet(title="Pastas de Rede")
            sheet_pastas.append([f.replace("_", " ").title() for f in pastas_fields])
            for row in pastas_data: sheet_pastas.append([row.get(f, '') for f in pastas_fields])
            for col in sheet_pastas.columns: sheet_pastas.column_dimensions[col[0].column_letter].width = 30
        
        # ABA 5: SOFTWARES
        softwares_fields = ['software', 'licenciado', 'acao', 'observacoes']
        softwares_data = parse_table_data(form, 'softwares', softwares_fields)
        if softwares_data:
            sheet_softwares = workbook.create_sheet(title="Softwares")
            sheet_softwares.append([f.replace("_", " ").title() for f in softwares_fields])
            for row in softwares_data: sheet_softwares.append([row.get(f, '') for f in softwares_fields])
            for col in sheet_softwares.columns: sheet_softwares.column_dimensions[col[0].column_letter].width = 30

        # ABA 6: EQUIPAMENTOS
        equipamentos_fields = ['tipo', 'modelo', 'patrimonio', 'sistema_operacional']
        equipamentos_data = parse_table_data(form, 'equipamentos', equipamentos_fields)
        if equipamentos_data:
            sheet_equip = workbook.create_sheet(title="Equipamentos")
            sheet_equip.append([f.replace("_", " ").title() for f in equipamentos_fields])
            for row in equipamentos_data: sheet_equip.append([row.get(f, '') for f in equipamentos_fields])
            for col in sheet_equip.columns: sheet_equip.column_dimensions[col[0].column_letter].width = 30
        
        # Aplica estilos a todas as abas
        for sheet in workbook.worksheets:
            if sheet.max_row > 0:
                for cell in sheet["1:1"]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                for row in sheet.iter_rows():
                    for cell in row: cell.border = thin_border
        
        # --- ETAPA 2: SALVAR O ARQUIVO EM VEZ DE ENVIAR PARA DOWNLOAD ---
        
        # Gera um nome de arquivo único para evitar sobreposição e facilitar a identificação
        nome_colaborador = form.get('nome_completo', 'colaborador').strip().replace(' ', '_')
        tipo_solicitacao = form.get('tipo_solicitacao', 'solicitacao').strip().replace(' ', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"{tipo_solicitacao}_{nome_colaborador}_{timestamp}.xlsx"
        
        # Constrói o caminho completo para salvar o arquivo
        caminho_completo_arquivo = os.path.join(PASTA_AUTOMACAO, nome_arquivo)
        
        # Salva o arquivo Excel no caminho especificado
        workbook.save(caminho_completo_arquivo)
        
        # --- ETAPA 3: FORNECER FEEDBACK AO USUÁRIO ---
        
        flash("Solicitação enviada com sucesso! O processo de automação foi iniciado.", "success")

    except Exception as e:
        # Captura qualquer erro durante o processo e informa ao usuário
        flash(f"Ocorreu um erro ao processar sua solicitação: {e}", "danger")
    
    # Redireciona o usuário de volta para a página do formulário em caso de sucesso ou falha
    # As mensagens de flash serão exibidas no topo da página (se o seu template suportar)
    return redirect(url_for('solicitacao'))


# --- ROTAS PLACEHOLDER (sem alteração) ---
@app.route('/aprovacoes')
@login_required
def aprovacoes():
    if getattr(current_user, 'role', 'user') != 'gestor':
        flash("Você não tem permissão para acessar esta página.", "warning")
        return redirect(url_for('solicitacao'))
    return render_template('base.html', page_content="<h1>Página de Aprovações</h1><p>Aqui o gestor poderá ver e aprovar as solicitações pendentes.</p>")

@app.route('/relatorios')
@login_required
def relatorios():
    return render_template('base.html', page_content="<h1>Página de Relatórios</h1><p>Relatórios de auditoria e conformidade estarão aqui.</p>")

@app.route('/configuracoes')
@login_required
def configuracoes():
    return render_template('base.html', page_content="<h1>Página de Configurações</h1><p>Configurações gerais do sistema.</p>")